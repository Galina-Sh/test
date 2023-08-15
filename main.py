import os
import tqdm
import openpyxl
import re
import psql_connection as psql
import datetime as dt
from nested_dict import nested_dict
import json

list_parametrs = ['aircraft_cabin_width', 'aircraft_engine_count', 'aircraft_engine_type', 'aircraft_height',
               'aircraft_jurisdiction', 'aircraft_length', 'aircraft_type_designator', 'aircraft_wingspan',
               'approach_noise_fact', 'approach_noise_limit', 'cost_article_id', 'cumulative_margin', 'document_date',
               'dp_foo_hc', 'fire_fighting_category', 'flight_in_area', 'flight_in_continental', 'flight_in_country',
               'flight_in_county', 'flight_in_distance', 'flight_in_EEA', 'flight_in_eu', 'flight_in_from_airport',
               'flight_in_holiday', 'flight_in_intercontinental', 'flight_in_locality', 'flight_in_region',
               'flight_in_schengen', 'flight_in_status', 'flight_in_technical', 'flight_in_transit', 'flight_out_area',
               'flight_out_continental', 'flight_out_country', 'flight_out_county', 'flight_out_distance',
               'flight_out_EEA', 'flight_out_eu', 'flight_out_holiday', 'flight_out_intercontinental',
               'flight_out_locality', 'flight_out_region', 'flight_out_schengen', 'flight_out_status',
               'flight_out_technical', 'flight_out_to_airport', 'flight_out_transit', 'flyover_noise_fact',
               'flyover_noise_limit', 'fuel_flow_approach', 'fuel_flow_climb_out', 'fuel_flow_idle',
               'fuel_flow_take_off', 'fuel_unit', 'fuel_uplift_date', 'fuel_volume', 'ic_id', 'is_budget', 'is_quotes',
               'jet_group', 'lateral_noise_fact', 'lateral_noise_limit', 'mglw', 'mtow', 'noise_chapter',
               'nox_approach', 'nox_climb_out', 'nox_idle', 'nox_take_off', 'nox_total', 'operator_jurisdiction', 'pax_in', 'pax_out',
               'prk_end_date', 'prk_end_time', 'prk_hours', 'prk_start_date', 'prk_start_time', 'quantity',
               'seat_capacity', 'service_type_id', 'service_type_name', 'vat_exists', 'icao_wtc', 'aircraft_width_type',
               'base_airport', 'crew_in', 'crew_out', 'flight_in_sunrise', 'flight_out_sunrise', 'flight_in_sunset', 'flight_out_sunset']

list_of_valid_functions = ['FLOOR', 'DATE', 'DATEDIF', 'DATEVALUE', 'DAY', 'DAYS', 'EDATE', 'EOMONTH', 'HOUR', 'MINUTE',
                           'MONTH', 'NOW', 'SECOND', 'TIME', 'TIMEVALUE', 'TODAY', 'WEEKDAY', 'YEAR', 'YEARFRAC',
                           'BIN2DEC', 'BIN2HEX', 'BIN2OCT', 'DEC2BIN', 'DEC2HEX', 'DEC2OCT', 'HEX2BIN', 'HEX2DEC', 'HEX2OCT',
                           'OCT2BIN', 'OCT2DEC', 'OCT2HEX', 'IRR', 'NPV', 'PMT', 'PV', 'SLN', 'VDB', 'XIRR', 'XNPV', 'ISBLANK',
                           'ISERR', 'ISERROR', 'ISEVEN', 'ISNA', 'ISNUMBER', 'ISODD', 'ISTEXT', 'NA', 'AND', 'IF',
                           'IFERROR', 'IFS', 'NOT', 'OR', 'XOR', 'COLUMN', 'COLUMNS', 'HLOOKUP', 'INDEX', 'INDIRECT',
                           'LOOKUP', 'MATCH', 'OFFSET', 'ROW', 'ROWS', 'VLOOKUP', 'ABS', 'CEILING', 'CEILING.MATH',
                           'CEILING.PRECISE', 'EVEN', 'FACT', 'FACTDOUBLE', 'CONVERT', 'TRANSPOSE', 'WORKDAY', 'SUBSTITUTE', 'SEARCH',
                           'TEXTJOIN', 'VALUE', 'UPPER', 'TRIM', 'RIGHT', 'REPLACE', 'MID', 'LOWER', 'LEN', 'LEFT', 'FIND', 'CONCATENATE',
                           'CONCAT', 'SMALL', 'MINIFS', 'MINA', 'MIN', 'MAXIFS', 'MAXA', 'MAX', 'LINEST', 'LARGE',
                           'COUNTIFS', 'COUNTIF', 'COUNTBLANK', 'COUNTA', 'COUNT', 'AVERAGEIFS', 'AVERAGEIF',
                           'AVERAGEA', 'AVERAGE', 'TRUNC', 'SUMPRODUCT', 'SUMIFS', 'SUMIF', 'SUM', 'SIGN', 'ROUNDUP', 'ROUNDDOWN', 'ROUND',
                           'RANDBETWEEN', 'RAND', 'POWER', 'ODD', 'MROUND', 'MOD', 'LOG10', 'LOG', 'LN', 'INT', 'FLOOR.PRECISE', 'FLOOR.MATH', 'MEDIAN', 'SQRT', 'EXACT']

validator_dict = nested_dict()

# Подключение к базам данных через вызов классов
myjet = psql.Myjet()
bench = psql.Bench()

path = '/Users/galina.tishchenkova/Desktop/Handling/'

for country in os.listdir(path):
    if country == ".DS_Store" or country == "Readme.md":
        continue
    for filename in tqdm.tqdm(os.listdir(f'{path}/{country}'), desc=country):
        if filename == ".DS_Store" or filename == "Readme.md":
            continue

        wb = openpyxl.load_workbook(f'{path}/{country}/{filename}', data_only=False)
        icao = filename[0:4]
        print(icao)
        wb_sheets = wb.sheetnames
        ws_params = wb['params']
        ws_main = wb['Main']
        ws_filter = wb['FILTERS_ADD_COORDS']
        main_handler_list = []
        main_handler_dict = {}
        handlers_columns_index = []
        cost_articles_list = []
        cost_ids = []
        services_list = []
        aircrafts = []
        authority = ['Eurocontrol', 'Austro Control', 'Aviation authority', 'FSUE State ATM Corporation']

        # Проверка: размер файла не превышает 2 МБ
        if os.path.getsize(f'{path}/{country}/{filename}') / 1024 / 1024 > 2:
            print('The file size cannot exceed 2 MB')
            validator_dict[f'{icao}']['check_1']['status']['error'] = 'The file size cannot exceed 2 MB'
        else:
            validator_dict[f'{icao}']['check_1']['status'] = 'ok'

        # Проверка: наименование файла соответствует маске шаблона
        if re.match(r'^[A-Z0-9]{4}\.xlsx$', filename) is None:
            print("The file name doesn't match the template mask")
            validator_dict[f'{icao}']['check_2']['status']['error'] = "The file name doesn't match the template mask"
        else:
            validator_dict[f'{icao}']['check_2']['status'] = 'ok'

        # Проверка: наличие обязательных листов в шаблоне
        expected_sheets = ['Main', 'params', 'AIP', 'FILTERS_ADD_COORDS']
        missing_sheets = [sheet for sheet in expected_sheets if sheet not in wb_sheets]  # соотносим имена листов в локации со списком expected_sheet_names и сохраняем недостающие в список missing_sheet_names
        if missing_sheets:
            print(f'Required sheet {str(missing_sheets)[1:-1]} missing')
            validator_dict[f'{icao}']['check_3.1']['status']['error'] = f'Required sheet {str(missing_sheets)[1:-1]} missing'
        else:
            validator_dict[f'{icao}']['check_3.1']['status'] = 'ok'

        # Проверка: корректное наименование DB-листов в шаблоне
        db_expected_names = ['DB_MTOW_RATES', 'DB_ATD_RATES', 'DB_MGLW_RATES', 'DB_AIRCRAFT_WINGSPAN_RATES',
                             'DB_SEAT_CAPACITY_RATES', 'DB_AIRCRAFT_LENGTH_RATES', 'DB_EXCEPTIONS', 'DB_FIXED_RATES']
        db_sheets = [re.sub(r'RATES.*', 'RATES', sheet) for sheet in wb_sheets if sheet.startswith('DB') and sheet not in expected_sheets]  # список db_sheet_names с листами DB из шаблона, 5 последних символов удалены
        missing_db_sheets = [sheet for sheet in db_sheets if sheet not in db_expected_names]
        if missing_db_sheets:
            print(f'Error in sheet title {str(missing_db_sheets)[1:-1]}')
            validator_dict[f'{icao}']['check_3.2']['status']['error'] = f'Error in sheet title {str(missing_db_sheets)[1:-1]}'
        else:
            validator_dict[f'{icao}']['check_3.2']['status'] = 'ok'

        # Далее проверки на Main
        for row in ws_main.iter_rows(max_row=1, max_col=ws_main.max_column):
            for cell in row:
                if cell.value is not None and '=' not in str(cell.value) and cell.value.strip() not in ('Cost article', 'Handlers', 'MAND', 'IGNO', 'QIGNO', 'FIGNO', 'RIGNO'):
                    col_index = cell.column
                    row_index = cell.row

                    handler_name = cell.value
                    main_handler_list.append(handler_name)
                    handlers_col_index = cell.column_letter
                    handlers_columns_index.append(handlers_col_index)
                    handler_id = ws_main.cell(row=row_index + 1, column=col_index).value
                    main_handler_dict[handler_id] = handler_name  # добавляем в словарь main_handler_dict значение id как ключ и значение name

                    # Проверка: наличие currency в шапке шаблона
                    currency = ws_main.cell(row=row_index + 4, column=col_index)
                    if currency.value is not None:
                        validator_dict[f'{icao}']['check_4.1']['status'][f'{currency}'] = 'ok'

                        if re.match(r'[A-Z]{3}$', currency.value) is None:
                            print(f'Invalid currency value in {currency}: {currency.value}')
                            validator_dict[f'{icao}']['check_4.2']['status']['error'][f'{currency}'] = f'Invalid currency value in {currency}: {currency.value}'
                        else:
                            validator_dict[f'{icao}']['check_4.2']['status'][f'{currency}'] = 'ok'

                    else:
                        print(f'Missing currency value in {currency}')
                        validator_dict[f'{icao}']['check_4.1']['status']['error'][f'{currency}'] = f'Missing currency value in {currency}'

                    # Проверка: дата начала действия поставщика не больше даты конца действия поставщика
                    period_start_date = ws_main.cell(row=row_index + 2, column=col_index)
                    period_end_date = ws_main.cell(row=row_index + 3, column=col_index)

                    if not isinstance(period_start_date.value, str):
                        print(f'Incorrect data format in {period_start_date}')
                        validator_dict[f'{icao}']['check_5.1']['status']['error'][f'{period_start_date}'] = f'Incorrect data format in {period_start_date}'
                    else:
                        validator_dict[f'{icao}']['check_5.1']['status'][f'{period_start_date}'] = 'ok'

                    if period_end_date.value is not None:
                        if isinstance(period_end_date.value, str):
                            validator_dict[f'{icao}']['check_5.2']['status'][f'{period_end_date}'] = 'ok'

                            period_start = dt.datetime.strptime(period_start_date.value, '%Y-%m-%d').date()
                            period_end = dt.datetime.strptime(period_end_date.value, '%Y-%m-%d').date()
                            if period_start >= period_end:
                                print(f'Date in {period_start_date} more than value in {period_end_date}')
                                validator_dict[f'{icao}']['check_5.3']['status']['error'][f'{period_end_date}'] = f'End date in {period_end_date} is less then start date in {period_start_date}'
                            else:
                                validator_dict[f'{icao}']['check_5.3']['status'][f'{period_end_date}'] = 'ok'

                        else:
                            print(f'Incorrect data format in {period_end_date}')
                            validator_dict[f'{icao}']['check_5.1']['status']['error'][f'{period_end_date}'] = f'Incorrect data format in {period_end_date}'

                    # Проверка: наличие MAND у Eurocontrol, Austrocontrol, Aviation Authority, FSUE State ATM Corporation
                    if handler_name in authority:
                        attribute = ws_main.cell(row=row_index, column=col_index + 1)
                        if attribute.value != 'MAND':
                            print(f'Provider {handler_name} is missing required attribute "MAND"')
                            validator_dict[f'{icao}']['check_6']['status']['error'][f'{attribute}'] = f'Provider \'{handler_name}\' is missing required attribute \'MAND\''
                        else:
                            validator_dict[f'{icao}']['check_6']['status'][f'{attribute}'] = 'ok'

        # Проверка: корректное наименование поставщика
        companies_request = myjet.select(f"SELECT id, name FROM companies WHERE id IN ({','.join(map(str, list(main_handler_dict.keys())))})")  # поиск в базе по id хендлера
        db_handler_dict = {row[0]: row[1] for row in companies_request}

        for id, name in main_handler_dict.items():
            if id in db_handler_dict:
                validator_dict[f'{icao}']['check_7.1']['status'][id] = 'ok'

                if db_handler_dict[id] != name:
                    print(f'Name \'{name}\' does not exist in "companies" database')
                    validator_dict[f'{icao}']['check_7.2']['status']['error'][name] = f'Name \'{name}\' does not exist in \'companies\' database'
                else:
                    validator_dict[f'{icao}']['check_7.2']['status'][name] = 'ok'

            else:
                print(f'Handler id \'{id}\' does not exist in "companies" database')
                validator_dict[f'{icao}']['check_7.1']['status']['error'][id] = f'Handler id \'{id}\' does not exist in \'companies\' database'

        # Проверка: корректное наименование cost article и их id в столбце ‘A’
        for row in ws_main.iter_rows(min_row=7, max_row=ws_main.max_row, max_col=1):  # составление списка из cost article в столбце А
            for cell in row:
                if cell.value is not None:
                    cost_article = cell.value
                    cost_articles_list.append(cost_article)

                    cost_id = cost_article.split('.')[0]
                    cost_ids.append(cost_id)

        cost_articles_dict = {int(i.split('.')[0]): i.split('.')[1].strip() for i in cost_articles_list}
        cost_article_request = bench.select(f"select id, name from cost_articles where id in ({','.join(map(str, cost_articles_dict))})")  # поиск в базе по id коста
        db_cost_articles = {row[0]: row[1] for row in cost_article_request}

        for id, name in cost_articles_dict.items():
            if id in db_cost_articles:
                validator_dict[f'{icao}']['check_8.1']['status'][id] = 'ok'

                if db_cost_articles[id] != name:
                    print(f'Cost article name \'{name}\' does not exist in "cost_articles" database')
                    validator_dict[f'{icao}']['check_8.2']['status']['error'][name] = f'Cost article name \'{name}\' does not exist in \'cost_articles\' database'
                else:
                    validator_dict[f'{icao}']['check_8.2']['status'][name] = 'ok'

            else:
                print(f'Cost article id \'{id}\' does not exist in "cost_articles" database')
                validator_dict[f'{icao}']['check_8.1']['status']['error'][id] = f'Cost article id \'{id}\' does not exist in \'cost_articles\' database'

        # Проверка: корректное наименование сервис типов в столбце ‘B’
        for row in ws_main.iter_rows(min_row=7, max_row=ws_main.max_row, min_col=2, max_col=2):
            for cell in row:
                if cell.value is not None and cell.value.strip().lower() not in ('disbursement', 'vat', 'sales tax'):
                    service_type = cell.value.strip()
                    escaped_value = service_type.replace("'", "''")  # зеркалим апостроф вторым рядом
                    services_list.append(escaped_value)

        unique_list = set(services_list)
        unique_list2 = str(unique_list).replace('"', "'")  # заменяем все двойные кавычки на одинарные

        type_of_services_request = bench.select(f"select service_type_name from type_of_services where service_type_name in ({str(unique_list2)[1:-1]})")
        db_type_of_services = [service[0] for service in type_of_services_request]
        error_service_type = [service for service in services_list if service.replace("''", "'") not in list(set(db_type_of_services))]

        if error_service_type:
            print(f'Service type name {str(error_service_type)[1:-1]} does not exist in "type_of_services" database')
            validator_dict[f'{icao}']['check_9']['status']['error'] = f'Service type name {str(error_service_type)[1:-1]} does not exist in \'type_of_services\' database'
        else:
            validator_dict[f'{icao}']['check_9']['status'] = 'ok'

        # Проверка: наличие 299 кост артикла в каждом аэропортовом шаблоне
        if '299' not in cost_ids:
            print('Required 299 cost article missing in template')
            validator_dict[f'{icao}']['check_10']['status']['error'] = 'Required 299 cost article missing in template'
        else:
            validator_dict[f'{icao}']['check_10']['status'] = 'ok'

        # Далее проверки по всем листам
        for sheet in wb.worksheets:
            for row in sheet:
                for cell in row:
                    # Все непустые ячейки формата string, содержащие формулу
                    if cell.value is not None and type(cell.value) is not None and isinstance(cell.value,str) and cell.value.startswith('='):

                        # Проверка: cсылка на другой файл Excel в ячейке
                        if '[' in cell.value and ']' in cell.value:
                            print(f'External link in {cell}')  # =[1]AIP!$A$1 внешняя ссылка в виде [1]
                            validator_dict[f'{icao}']['check_11']['status']['error'][f'{cell}'] = f'External link in {cell}'
                        else:
                            validator_dict[f'{icao}']['check_11']['status'][f'{cell}'] = 'ok'

                        # Проверка на наличие функции в книге, которая не поддерживается Pycell
                        if re.search(r'[A-Z]{1,}\(', cell.value) is not None:
                            function = re.findall(r'[A-Z]{1,}\(', cell.value)  # список из всех функций соответствующих паттерну в непустой ячейке
                            function_list = [re.sub(r'\(', '', func) for func in function]
                            not_valid_function_list = [func for func in function_list if func not in list_of_valid_functions]
                            if not_valid_function_list:
                                print(f'Not valid function {str(not_valid_function_list)[1:-1]} in {cell}')
                                validator_dict[f'{icao}']['check_12']['status']['error'][f'{cell}'] = 'Not valid function {str(not_valid_function_list)[1:-1]} in {cell}'
                            else:
                                validator_dict[f'{icao}']['check_12']['status'][f'{cell}'] = 'ok'

        # Далее проверка формата ячеек на листе Main
        handler_columns = '(' + "'" + "', '".join(
            handlers_columns_index) + "'" + ')'  # список индексов столбцов с хендлерами в виде ('A', 'B', 'C')
        # в столбце 'B' находим все строки с Disbursement/VAT/Sales tax, затем строки с сервис типами, затем пустые строки
        for col in ws_main.iter_cols(min_row=6, max_row=ws_main.max_row, min_col=2, max_col=2):
            for cell in col:

                # Проверка: ячейка `disbursement`, `VAT`, `Sales tax` - `percentage`
                if cell.value is not None and cell.value.strip().lower() in ('disbursement', 'vat', 'sales tax'):
                    row_index_percentage = cell.row
                    # проверяем формат ячеек в найденных строках
                    for row in ws_main.iter_rows(min_row=6, max_row=ws_main.max_row, min_col=3, max_col=ws_main.max_column):
                        for cell in row:
                            col_index = cell.column_letter
                            if col_index in handler_columns:
                                if cell.row == row_index_percentage:
                                    if cell.value is not None and cell.number_format.endswith('%') == False:  # cell.number_format: ('0%','0.0%', '0.00%', '0.000%')
                                        print(f'Wrong cell format in {cell} (percentage)')
                                        validator_dict[f'{icao}']['check_13']['status']['error'][f'{cell}'] = f'Wrong cell format in {cell} (percentage)'
                                    else:
                                        validator_dict[f'{icao}']['check_13']['status'][f'{cell}'] = 'ok'

                # Проверка: ячейки с суммами по сервис типам в формате - currency
                if cell.value is not None and cell.value.strip().lower() not in ('disbursement', 'vat', 'sales tax'):
                    row_index_service = cell.row
                    # проверяем формат ячеек в найденных строках
                    for row in ws_main.iter_rows(min_row=6, max_row=ws_main.max_row, min_col=3, max_col=ws_main.max_column):
                        for cell in row:
                            col_index = cell.column_letter
                            if col_index in handler_columns:
                                if cell.value is not None:  # добавляем условие is not None, т.к. конкретного севис типа у конкретного поставщика может не быть (пустая ячейка)
                                    if cell.row == row_index_service:
                                        if '$' not in cell.number_format:  # cell.style != 'Currency' не подходит
                                            print(f'Wrong cell format in {cell} (currency)')
                                            validator_dict[f'{icao}']['check_14']['status']['error'][f'{cell}'] = f'Wrong cell format in {cell} (currency)'
                                        else:
                                            validator_dict[f'{icao}']['check_14']['status'][f'{cell}'] = 'ok'

            # Проверка: пустые ячейки между кост артиклами в формате - `general`
            for row in ws_main.iter_rows(min_row=6, max_row=ws_main.max_row, min_col=3, max_col=ws_main.max_column):
                for cell in row:
                    col_index = cell.column_letter
                    if col_index in handler_columns:
                        if cell.value is None:
                            if cell.number_format != 'General':
                                print(f'Wrong cell format in {cell} (general)')
                                validator_dict[f'{icao}']['check_15.1']['status']['error'][f'{cell}'] = f'Wrong cell format in {cell} (general)'
                            else:
                                validator_dict[f'{icao}']['check_15.1']['status'][f'{cell}'] = 'ok'

        # Проверка: ячейки между поставщиками (где атрибуты) + столбцы A, B в формате - `general`
        for col in ws_main.iter_cols(max_row=ws_main.max_row, max_col=ws_main.max_column):
            for cell in col:
                col_index = cell.column_letter
                if col_index not in handler_columns:
                    if cell.number_format != 'General':
                        print(f'Wrong cell format in {cell} (general)')
                        validator_dict[f'{icao}']['check_15.2']['status']['error'][f'{cell}'] = f'Wrong cell format in {cell} (general)'
                    else:
                        validator_dict[f'{icao}']['check_15.2']['status'][f'{cell}'] = 'ok'

        # Проверка: корректное наименование атрибуций сервисов
        for i, column in enumerate(ws_main.iter_cols(min_row=7, min_col=3, max_col=ws_main.max_column), start=1):
            if i % 2 == 0:
                for cell in column:
                    if cell.value is not None and isinstance(cell.value, str) and not cell.value.startswith('='):
                        if cell.value.strip() not in ('BAS', 'BHS', 'EXTRA', 'SKY'):
                            print(f'Incorrect attribute "{cell.value}" in {cell}')
                            validator_dict[f'{icao}']['check_16']['status']['error'][f'{cell}'] = f'Incorrect attribute \'{cell.value}\' in {cell}'
                        else:
                            validator_dict[f'{icao}']['check_16']['status'][f'{cell}'] = 'ok'

        # Проверка: cервисы в столбце ‘B’ на листе params должны соответствовать наименованию сервисов в столбце ‘B’ на листе Main
        list_main_services = []
        list_params_services = []

        for row in ws_main.iter_rows(min_row=7, max_row=ws_main.max_row, min_col=2, max_col=2):
            for cell in row:
                if cell.value is not None and cell.value.strip().lower() not in ('disbursement', 'vat', 'sales tax'):
                    list_main_services.append(cell.value)

        for row in ws_params.iter_rows(min_row=2, max_row=ws_params.max_row, min_col=2, max_col=2):
            for cell in row:
                if cell.value is not None:
                    list_params_services.append(cell.value)

        list_main = [x for x in list_main_services if x not in list_params_services]
        if list_main:
            print(f'Service {str(list_main)[1:-1]} not on \'params\' sheet')
            validator_dict[f'{icao}']['check_17.1']['status']['error'] = f'Service {str(list_main)[1:-1]} not on \'params\' sheet'
        else:
            validator_dict[f'{icao}']['check_17.1']['status'] = 'ok'

        list_params = [x for x in list_params_services if x not in list_main_services]
        if list_params:
            print(f'Service {str(list_params)[1:-1]} not on \'Main\' sheet')
            validator_dict[f'{icao}']['check_17.2']['status']['error'] = f'Service {str(list_params)[1:-1]} not on \'Main\' sheet'
        else:
            validator_dict[f'{icao}']['check_17.2']['status'] = 'ok'

        # Проверка: корректное написание всех параметров
        for row in ws_params.iter_rows(max_row=1, min_col=5, max_col=ws_params.max_column):
            for cell in row:
                if cell.value is not None:
                    if cell.value not in list_parametrs:
                        print(f'Parameter {cell.value} in {cell} not in params meta')
                        validator_dict[f'{icao}']['check_18']['status']['error'][f'{cell}'] = f'Parameter {cell.value} in {cell} not in params meta'
                    else:
                        validator_dict[f'{icao}']['check_18']['status'][f'{cell}'] = 'ok'

        named_range = wb.defined_names.definedName  # wb.defined_names.definedName работает только с openpyxl <= 3.0.10
        # Проверка: параметры aircraft_engine_model, aircraft_acoustic_group, acoustic_group, flight_hours, flight_count должны отсутствовать в диспетчере имён
        for name in named_range:
            if name.name in ('aircraft_engine_model', 'aircraft_acoustic_group', 'acoustic_group', 'flight_hours', 'flight_count'):
                print(f'Unexpected name {name.name} in Name Manager')
                validator_dict[f'{icao}']['check_19.1']['status']['error'][f'{name.name}'] = f'Unexpected name {name.name} in Name Manager'
            else:
                validator_dict[f'{icao}']['check_19.1']['status'][f'{name.name}'] = 'ok'

        # Проверка: cсылка на другой файл Excel в диспетчере имён
        for name in named_range:
            if '[' in name.value or ']' in name.value:
                print(f'External link in value for name {name.name} in Name Manager')
                validator_dict[f'{icao}']['check_19.2']['status']['error'][f'{name.value}'] = f'External link in value for name {name.name} in Name Manager'
            else:
                validator_dict[f'{icao}']['check_19.2']['status'][f'{name.value}'] = 'ok'

        # Проверка: в диспетчере имен не должно быть ошибок REF!
        for name in named_range:
            if '#REF' in name.value:
                print(f'\'#REF\' value for name {name.name} in Name Manager')
                validator_dict[f'{icao}']['check_20']['status']['error'][f'{name.name}'] = f'\'#REF\' value for name {name.name} in Name Manager'
            else:
                validator_dict[f'{icao}']['check_20']['status'][f'{name.name}'] = 'ok'

        # Далее проверки на листах DB
        for sheet in wb_sheets:
            if 'DB' in sheet:
                db_sheet = wb[sheet]

                # Проверка: отсутствие формул на листах DB_()_RATES
                for row in db_sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None and isinstance(cell.value, str):
                            if cell.value.startswith('='):
                                print(f'Unexpected formula on {db_sheet} in {cell}')
                                validator_dict[f'{icao}']['check_21']['status']['error'][f'{cell}'] = f'Unexpected formula on {db_sheet} in {cell}'
                            else:
                                validator_dict[f'{icao}']['check_21']['status'][f'{cell}'] = 'ok'

                # Проверка: корректное одинаковое наименование поставщиков на всех листах (Main, DB_()_RATES)
                for row in db_sheet.iter_rows(max_row=1, min_col=1, max_col=db_sheet.max_column):
                    for cell in row:
                        if cell.value is not None and cell.value != 'Airport' and cell.value not in main_handler_list:
                            print(f'{cell.value} in {cell} is not in Handlers list')
                            validator_dict[f'{icao}']['check_22']['status']['error'][f'{cell}'] = f'{cell.value} in {cell} is not in Handlers list'
                        else:
                            validator_dict[f'{icao}']['check_22']['status'][f'{cell}'] = 'ok'

                # Проверка: корректное наименование моделей самолетов на db_atd
                if 'ATD' in sheet:
                    db_atd_sheet = wb[sheet]
                    for col in db_atd_sheet.iter_cols(min_row=3, max_row=db_atd_sheet.max_row, max_col=1):
                        for cell in col:
                            if cell.value is not None and type(cell.value) is not None and isinstance(cell.value, str):
                                if cell.value not in aircrafts:
                                    aircrafts.append(cell.value.strip())

        aircraft_models_request = myjet.select("select type_designator from aircraft_models where type_designator is not null;")
        aircraft_models_list = [item[0] for item in aircraft_models_request]
        invalid_aircrafts = [i for i in aircrafts if i not in aircraft_models_list]
        if invalid_aircrafts:
            print(f'Non-existent aircraft model {str(invalid_aircrafts)[1:-1]} on DB_ATD_RATES-sheet')
            validator_dict[f'{icao}']['check_23']['status']['error'] = f'Non-existent aircraft model {str(invalid_aircrafts)[1:-1]} on DB_ATD_RATES-sheet'
        else:
            validator_dict[f'{icao}']['check_23']['status'] = 'ok'


        for sheet in wb_sheets:
            db_match = re.compile(r'(DB_[A-Z_]+_)\d{4}')
            if db_match.match(sheet) is not None:

                # Проверка: наличие наименования сервиса во второй строке, если есть запись в первой строке на вкладке DB
                db_rates = wb[sheet]
                for col in range(2, db_rates.max_column + 1):
                    col_data = [db_rates.cell(row=row, column=col).value for row in range(2, db_rates.max_row + 1)]  # список, содержащий значения ячеек в столбце
                    if any(col_data) and col_data[0] is None:  # пропускаем столбцы, в которых все значения равны `None` (полностью пустой столбец) и выбираем столбцы где первое значение `None` (пустая ячейка в начале столбца)
                        print(f'Service name is missing on {str(db_rates)[1:-1]} in column {col}')
                        validator_dict[f'{icao}']['check_24']['status']['error'][f'{db_rates}'][f'{col}'] = f'Service name is missing on {str(db_rates)[1:-1]} in column {col}'
                    else:
                        validator_dict[f'{icao}']['check_24']['status'][f'{db_rates}'][f'{col}'] = 'ok'


        # Проверка: наличие вкладок с текущим годом для листов db_()_rates
        sheets_to_upd = []

        for sheet in wb_sheets:
            counter = 0
            db_match = re.compile(r'(DB_[A-Z_]+_)\d{4}')
            if db_match.match(sheet) is not None:

                # Проверка: ставки на последней вкладке листа db_()_rates должны быть обновлены
                last_sheet = f'{sheet[:-4]}{dt.datetime.now().year}'
                if last_sheet in wb_sheets:
                    validator_dict[f'{icao}']['check_25']['status'][f'{last_sheet}'] = 'ok'

                    previous_sheet = f'{sheet[:-4]}{dt.datetime.now().year - 1}'
                    if previous_sheet in wb_sheets:  # сравниваем ставки за текущий год со ставками за предыдущий год
                        counter = 0
                        for row1, row2 in zip(wb[last_sheet].iter_rows(max_row=wb[last_sheet].max_row,
                                                                       max_col=wb[last_sheet].max_column,
                                                                       values_only=True),
                                              wb[previous_sheet].iter_rows(max_row=wb[previous_sheet].max_row,
                                                                           max_col=wb[previous_sheet].max_column,
                                                                           values_only=True)):
                            for cell1, cell2 in zip(row1, row2):
                                if cell1 is not None and cell2 is not None:
                                    if cell1 != cell2:
                                        counter += 1

                        if counter == 0:
                            sheets_to_upd.append(last_sheet)
                    else:
                        continue
                else:
                    print(f'Sheet {sheet[:-5]} has no current year')
                    validator_dict[f'{icao}']['check_25']['status']['error'][f'{last_sheet}'] = f'Sheet {sheet[:-5]} has no current year'

        if set(sheets_to_upd):
            print(f'Rates need to be updated on sheet {str(set(sheets_to_upd))[1:-1]}')
            validator_dict[f'{icao}']['check_26']['status']['error'] = 'Rates need to be updated on sheet {str(set(sheets_to_upd))[1:-1]}'
        else:
            validator_dict[f'{icao}']['check_26']['status'] = 'ok'


        if 'DB_FIXED_RATES' in wb_sheets:
            ws_fixed = wb['DB_FIXED_RATES']

            # Проверка: корректный формат даты в столбце ‘A’ на db_fixed
            for row in range(3, ws_fixed.max_row + 1):
                cell_value = ws_fixed['A' + str(row)].value
                if cell_value is not None:
                    if isinstance(cell_value, dt.datetime):
                        validator_dict[f'{icao}']['check_27.1']['status'][f'{cell_value}'] = 'ok'

                        date_str = cell_value.strftime("%d.%m.%Y")
                        if cell_value != dt.datetime.strptime(date_str, "%d.%m.%Y"):
                            print(f'Incorrect data format in row {row} on sheet DB_FIXED_RATES')
                            validator_dict[f'{icao}']['check_27.2']['status'][f'{cell_value}'] = f'Incorrect data format in row {row} on sheet DB_FIXED_RATES'
                        else:
                            validator_dict[f'{icao}']['check_27.2']['status'][f'{cell_value}'] = 'ok'

                    else:
                        print(f'Incorrect data format in row {row} on sheet DB_FIXED_RATES: {cell_value}')
                        validator_dict[f'{icao}']['check_27.1']['status'][f'{cell_value}'] = f'Incorrect data format in row {row} on sheet DB_FIXED_RATES: {cell_value}'

            # Проверка: наличие пустых ячеек на db_fixed
            if ws_fixed['A2'].value != None:
                global first_column_data
                global column_data
                validator_dict[f'{icao}']['check_28.1']['status'] = 'ok'

                for col in range(1, 2):
                    first_column_data = [ws_fixed.cell(row=row, column=col).value for row in
                                         range(2, ws_fixed.max_row + 1) if ws_fixed.cell(row=row, column=col).value is not None]  # список, содержащий значения ячеек в первом столбце

                for col in range(2, ws_fixed.max_column + 1):
                    column_data = [ws_fixed.cell(row=row, column=col).value for row in range(2, ws_fixed.max_row + 1) if
                                   ws_fixed.cell(row=row, column=col).value is not None]  # список, содержащий значения ячеек в столбце

                    if len(column_data) != len(first_column_data):
                        print(f'There must be no empty cells on sheet DB_FIXED_RATES')
                        validator_dict[f'{icao}']['check_28.2']['status']['error'][f'column {col}'] = f'There must be no empty cells on sheet DB_FIXED_RATES'
                    else:
                        validator_dict[f'{icao}']['check_28.2']['status'][f'column {col}'] = 'ok'

            else:
                print(f"Cell {ws_fixed['A2']} must be not None")
                validator_dict[f'{icao}']['check_28.1']['status']['error'] = f"Cell {ws_fixed['A2']} must be not None"

            # Проверка: выводить уведомление если ставки в df_fixed_rates из последнего периода соответствуют ставкам из предыдущего периода не менее чем на 75%.
            rates = []
            count = 0  # счетчик одинаковых значений
            last_row = ws_fixed.max_row
            global last_date_dt
            global last_rate

            for row in range(3, ws_fixed.max_row + 1):
                cell = ws_fixed['A' + str(row)]
                if cell.value is not None:
                    last_date = str(ws_fixed.cell(row=row, column=1).value)
                    last_date_dt = dt.datetime.strptime(last_date[:10], '%Y-%m-%d').date()
                    if int(last_date_dt.year) <= int(dt.date.today().year):  # берем значение последней ячейки, которая не больше нынешнего года
                        last_row = cell.row

            for col in ws_fixed.iter_cols(min_col=2, min_row=last_row, max_row=last_row, max_col=ws_fixed.max_column):  # берем значения из последней строки с крайней датой
                for cell in col:
                    if cell is not None:
                        col_index = cell.column
                        row_index = cell.row
                        rates.append(cell)
                        if row_index == 3:
                            count = 0
                        else:
                            last_rate = ws_fixed.cell(row=row_index - 1, column=col_index)  # значение в строке с предпоследней датой ставок

                    if cell.value == last_rate.value:
                        count += 1

            if last_date_dt < dt.date.today():
                print(f'Rates on DB_FIXED_RATES needs to be updated for current year')
                validator_dict[f'{icao}']['check_29.1']['status']['error'] = f'Rates on DB_FIXED_RATES sheet need to be updated for current year'
            else:
                validator_dict[f'{icao}']['check_29.1']['status'] = 'ok'

            if len(rates) != 0 and count / len(rates) >= 0.75:
                print(f'Sheet DB_FIXED_RATES has more than 75% of same rates in the last period')
                validator_dict[f'{icao}']['check_29.2']['status']['error'] = f'Sheet DB_FIXED_RATES has more than 75% of same rates in the last period'
            else:
                validator_dict[f'{icao}']['check_29.2']['status'] = 'ok'

        # Проверка: корректное наименование фильтров на вкладке FILTERS_ADD_COORDS
        all_filters = ['filter_by_aircraft_jet_group', 'filter_by_aircraft_type_designator', 'filter_by_aircraft_mtow',
                       'filter_by_aircraft_mglw', 'filter_by_aircraft_wingspan', 'filter_by_aircraft_length',
                       'filter_by_aircraft_seat_capacity']

        if 'FILTERS_ADD_COORDS' in wb_sheets:  # проверим столбцы C на листе FILTERS_ADD_COORDS и создадим список filters с уникальными значениями этого столбца
            validator_dict[f'{icao}']['check_30.1']['status'] = 'ok'

            filters = set(row[2] for row in ws_filter.iter_rows(min_row=2, values_only=True) if row[2])
            wrong_filters = [filter for filter in filters if filter not in all_filters]
            if wrong_filters:
                print(f'Error in filter title {str(wrong_filters)[1:-1]}')
                validator_dict[f'{icao}']['check_30.2']['status']['error'] = f'Error in filter title {str(wrong_filters)[1:-1]}'
            else:
                validator_dict[f'{icao}']['check_30.2']['status'] = 'ok'

        else:
            print("Required sheet 'FILTERS_ADD_COORDS' is missing")
            validator_dict[f'{icao}']['check_30.1']['status']['error'] = "Required sheet 'FILTERS_ADD_COORDS' is missing"

        # Проверка: ссылки на вкладки DB (кроме FIXED) в колонке ‘A’
        db_expected_names = ['DB_MTOW_RATES', 'DB_ATD_RATES', 'DB_MGLW_RATES', 'DB_AIRCRAFT_WINGSPAN_RATES',
                             'DB_SEAT_CAPACITY_RATES', 'DB_AIRCRAFT_LENGTH_RATES', 'DB_EXCEPTIONS']

        values_list = [cell.value.split('!')[0] for cell in ws_filter['A'] if cell.value is not None]  # список с названиями вкладок DB из столбца A листа FILTERS_ADD_COORDS
        wrong_values = [value for value in values_list if value not in db_expected_names]  # значения в столбце А листа FILTERS_ADD_COORDS не соответствуют названиям вкладок DB
        if wrong_values:
            print(f"Errors on sheet FILTERS_ADD_COORDS in column 'A': {wrong_values}")
            validator_dict[f'{icao}']['check_31']['status']['error'] = f"Errors on sheet FILTERS_ADD_COORDS in column 'A': {wrong_values}"
        else:
            validator_dict[f'{icao}']['check_31']['status'] = 'ok'

        value_counts = {}
        for value in values_list:
            if value in value_counts:
                value_counts[value] += 1
            else:
                value_counts[value] = 1

        for value, count in value_counts.items():  # количество значений из values_list каждого типа сохраняем в соотв. переменные
            globals()[f'count_{value}'] = count

        # Проверка на совпадение количества строк в столбце A на листе FILTERS_ADD_COORDS и количества столбцов на соответствующих вкладках:
        for name in db_expected_names:
            if name in value_counts:
                value_count = value_counts[name]
                sheet_names = [sheet_name for sheet_name in wb.sheetnames if re.match(rf'{name}_20\d+', sheet_name)]
                for sheet_name in sheet_names:
                    sheet = wb[sheet_name]
                    column_count = sum(1 for column in sheet.iter_cols(min_row=2, max_row=2) if any(cell.value for cell in column))
                    if value_count != column_count - 1:
                        print(f'Number of columns on sheet {sheet_name} does not match the number of filter rows on sheet FILTERS_ADD_COORDS')
                        validator_dict[f'{icao}']['check_32.1']['status']['error'][f'{sheet_name}'] = f'Number of columns on sheet {sheet_name} does not match the number of filter rows on sheet FILTERS_ADD_COORDS'
                    else:
                        validator_dict[f'{icao}']['check_32.1']['status'][f'{sheet_name}'] = 'ok'

        # Проверка на совпадение количества столбцов на вкладках с одним типом ставок:
        for name in db_expected_names:
            sheet_names = [sheet_name for sheet_name in wb.sheetnames if re.match(rf'{name}_20\d+', sheet_name)]
            column_counts = []
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                column_count = sum(1 for column in sheet.iter_cols(min_row=2, max_row=2) if any(cell.value for cell in column))
                column_counts.append(column_count - 1)
            if len(set(column_counts)) > 1:
                print(f'Number of columns on sheets {name} does not match')
                validator_dict[f'{icao}']['check_32.2']['status']['error'][f'{name}'] = f'Number of columns on sheets {name} does not match'
            else:
                validator_dict[f'{icao}']['check_32.2']['status'][f'{name}'] = 'ok'

# out_file = open("validator.json", "w")
# json.dump(validator_dict, out_file, indent=4, separators=(',', ':'))
# out_file.close()
#
# icao_list = []
#
# errors_dict = nested_dict()
# for key, value in validator_dict.items():
#     for sub_key, sub_value in value.items():
#         for sub1_key, sub1_value in sub_value.items():
#             if sub1_value == 'ok':
#                 continue
#             for sub2_key, sub2_value in sub1_value.items():
#                 if sub2_key == 'error':
#                     icao_list.append(key)
#                     errors_dict[key][sub_key][sub1_key][sub2_key] = sub2_value
#
# out_file = open("errors.json", "w")
# json.dump(errors_dict, out_file, indent=4, separators=(',', ':'))
# out_file.close()
#
# print(set(icao_list))
